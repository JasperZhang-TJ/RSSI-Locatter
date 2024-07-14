import pandas as pd
import numpy as np
from processExcel import main as preprocess_main
import time
import json
import subprocess
from openpyxl import load_workbook
import os

#实时把数据写入Excel的函数
def write_to_csv(filename, data, columns):
    file_exists = os.path.isfile(filename)
    df = pd.DataFrame(data, columns=columns)
    if not file_exists:
        df.to_csv(filename, mode='a', index=False, header=True)
    else:
        df.to_csv(filename, mode='a', index=False, header=False)

def write_to_excel(filename, sheet_name, data):
    try:
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = load_workbook(filename)
        if sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        else:
            startrow = 0
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
        writer.save()
    except FileNotFoundError:
        df = pd.DataFrame(data)
        df.to_excel(filename, sheet_name=sheet_name, index=False)

# 这个函数用来补全不全的receiverID
def generate_receiver_ids(start, end):
    start_num = int(start[1:])
    end_num = int(end[1:])
    return [f'D{str(i).zfill(2)}' for i in range(start_num, end_num + 1)]


# 原先用来储存表头的函数 现在没有调用
def save_ids_and_mappings(data, output_file):
    # 创建 TRANSMITTER_ID 映射字典
    transmitter_id_map = {41: 1, 42: 2, 43: 3, 44: 4, 45: 5, 46: 6, 47: 7, 48: 8, 49: 9}

    # 应用映射字典替换 TRANSMITTER_ID
    data['TRANSMITTER_ID'] = data['TRANSMITTER_ID'].replace(transmitter_id_map)

    # 获取所有的 RECEIVER_ID 和映射后的 TRANSMITTER_ID
    receiver_ids = data['RECEIVER_ID'].unique()
    transmitter_ids = data['TRANSMITTER_ID'].unique()

    # 创建一个字典来存储这些编号和映射关系
    ids_dict = {
        'RECEIVER_ID': sorted(receiver_ids),
        'TRANSMITTER_ID': sorted(transmitter_ids),
        'TRANSMITTER_MAPPING': transmitter_id_map
    }

    # 将这些编号和映射关系存储到一个 Excel 文件中
    with pd.ExcelWriter(output_file) as writer:
        # 创建 RECEIVER_ID 和 TRANSMITTER_ID 的 DataFrame
        ids_df = pd.DataFrame({
            'RECEIVER_ID': pd.Series(sorted(receiver_ids)),
            'TRANSMITTER_ID': pd.Series(sorted(transmitter_ids))
        })
        ids_df.to_excel(writer, sheet_name='IDs', index=False)

        # 创建 TRANSMITTER_ID 映射关系的 DataFrame
        mapping_df = pd.DataFrame(list(transmitter_id_map.items()),
                                  columns=['Original_TRANSMITTER_ID', 'Mapped_TRANSMITTER_ID'])
        mapping_df.to_excel(writer, sheet_name='Mapping', index=False)

    print("RECEIVER_ID and TRANSMITTER_ID along with their mappings have been saved successfully.")

# 初始化读取表格 并且生成 combinations.csv
def read_normalized_data_and_find_combinations(normalized_file, output_file, receiver_fill):
    # 读取 Excel 文件
    excel_data = pd.ExcelFile(normalized_file)

    all_combinations = []
    receiver_ids = []

    # 遍历每个 sheet
    for sheet_name in excel_data.sheet_names:
        # 读取 sheet 数据
        sheet_data = pd.read_excel(normalized_file, sheet_name=sheet_name, index_col=0)

        # 取出第一行
        first_row = sheet_data.iloc[0]
        receiver_ids = sheet_data.columns.tolist()

        # 打印第一列对应的时间
        timestamp = first_row.name
        print(f"Timestamp: {timestamp}")

        # 找出有值的 TRANSMITTER 和 RECEIVER 组合
        transmitter_id = sheet_name.split('_')[1]  # 从 sheet 名字提取 transmitter_id
        for receiver_id, value in first_row.items():
            if not pd.isna(value):
                all_combinations.append((timestamp, receiver_id, transmitter_id, value))

    # 补全 receiver_ids
    if receiver_fill and receiver_ids:
        min_receiver = "D00"
        max_receiver = max(receiver_ids)
        receiver_ids = generate_receiver_ids(min_receiver, max_receiver)

    # 创建一个 DataFrame 存储所有的组合
    # 先创建包含所有 Receiver 列的空 DataFrame
    combinations_df = pd.DataFrame(columns=['Timestamp'] + receiver_ids)
    combinations_df.set_index('Timestamp', inplace=True)

    # 对 RSSI 值从大到小排序
    all_combinations.sort(key=lambda x: x[3], reverse=True)

    # 找到最符合的组合并打印
    used_transmitters = set()
    used_receivers = set()
    selected_combinations = []
    for timestamp, receiver_id, transmitter_id, rssi in all_combinations:
        if transmitter_id not in used_transmitters and receiver_id not in used_receivers:
            selected_combinations.append((timestamp, receiver_id, transmitter_id, rssi))
            used_transmitters.add(transmitter_id)
            used_receivers.add(receiver_id)

    # 将组合数据填入 DataFrame 中
    for timestamp, receiver_id, transmitter_id, rssi in selected_combinations:
        if timestamp not in combinations_df.index:
            combinations_df.loc[timestamp] = [np.nan] * len(combinations_df.columns)
        combinations_df.at[timestamp, receiver_id] = transmitter_id

    # 补全缺失的 receiver 列
    if receiver_fill and receiver_ids:
        for receiver_id in receiver_ids:
            if receiver_id not in combinations_df.columns:
                combinations_df[receiver_id] = np.nan

    combinations_df = combinations_df.sort_index(axis=1)

    # 将透视表存储到一个新的 CSV 文件中
    combinations_df.to_csv(output_file)

    print(f"All combinations have been saved to {output_file}.")

# 读取 normalized_file 的第 k 行数据和 combinations_file 的前 k-1 行数据
def read_kth_row_and_combinations(normalized_file, combinations_file, k):
    """
    读取 normalized_file 的第 k 行数据和 combinations_file 的前 k-1 行数据。

    输入:
    - normalized_file: 归一化后的 Excel 文件名
    - combinations_file: 组合数据的 CSV 文件名
    - k: 要读取的行数 (从1开始)

    输出:
    - normalized_data: 包含第 k 行数据的列表 [(sheet_name, kth_row)]
    - combinations_data: 前 k-1 行的组合数据 DataFrame
    """
    # 读取 normalized_file 的第 k 行数据
    sheet_names = pd.ExcelFile(normalized_file).sheet_names
    normalized_data = []

    for sheet_name in sheet_names:
        sheet_data = pd.read_excel(normalized_file, sheet_name=sheet_name, index_col=0)
        if k <= len(sheet_data):
            kth_row = sheet_data.iloc[k - 1]
            normalized_data.append((sheet_name, kth_row))

    # 打印第 k 行数据的时间戳
    print(f"DataNormalized - Row {k}:")
    for sheet_name, row in normalized_data:
        timestamp = row.name
        print(f"Timestamp: {timestamp}")

    # 读取 combinations_file 的前 k-1 行数据
    combinations_df = pd.read_csv(combinations_file, index_col=0)
    if k - 1 <= len(combinations_df):
        combinations_data = combinations_df.iloc[:k - 1]
    else:
        combinations_data = combinations_df

    return normalized_data, combinations_data

# 假设我们定义了一个新的指标计算函数
def calculate_new_metric(receiver_id, transmitter_id, rssi, combinations_data, k):
    """
    计算新的指标，根据 RSSI 和接收器位置得分。

    输入:
    - receiver_id: 当前接收器的ID (str)
    - transmitter_id: 当前发射器的ID (str)
    - rssi: 接收到的信号强度指示 (float)
    - combinations_data: 包含之前组合数据的 DataFrame
    - k: 当前的行数 (int)

    输出:
    - 计算得到的指标 (float)
    """
    # 定义位置得分
    score_map = {
        0: 6,
        1: 3,
        -1: 5,
        2: 1,
        -2: 2,
    }

    def get_score(current_position, previous_position, num_receivers):
        diff = -(current_position - previous_position)
        if diff in score_map:
            return score_map[diff]
        else:
            return 0

    # 获取当前接收器的位置索引
    num_receivers = len(combinations_data.columns)
    if receiver_id not in combinations_data.columns:
        raise KeyError(f"Receiver ID {receiver_id} not found in combinations_data columns")
    current_position = combinations_data.columns.get_loc(receiver_id)

    total_score = 0

    # 遍历之前的所有行数据
    for idx, (_, row) in enumerate(combinations_data.iterrows()):
        decay_factor = 0.5 ** (k - idx - 1)
        for prev_receiver_id, prev_transmitter_id in row.items():
            if prev_transmitter_id == transmitter_id:
                previous_position = combinations_data.columns.get_loc(prev_receiver_id)
                total_score += get_score(current_position, previous_position, num_receivers) * decay_factor

    # 最终的指标计算
    new_metric = rssi * total_score
    return new_metric


# 根据新的指标分配 transmitters 到 receivers
def allocate_transmitters_based_on_new_metric(normalized_data, combinations_data,k):
    """
    根据新的指标分配 transmitters 到 receivers。

    输入:
    - normalized_data: 包含第 k 行数据的列表 [(sheet_name, kth_row)]
    - combinations_data: 前 k-1 行的组合数据 DataFrame

    输出:
    - selected_combinations: 分配后的组合列表 [(timestamp, receiver_id, transmitter_id, rssi)]
    """
    all_combinations = []

    for sheet_name, kth_row in normalized_data:
        timestamp = kth_row.name

        for receiver_id, value in kth_row.items():
            if not pd.isna(value):
                all_combinations.append((timestamp, receiver_id, sheet_name.split('_')[1], value))

    # 计算每个组合的新指标
    new_combinations = [
        (timestamp, receiver_id, transmitter_id, rssi, calculate_new_metric(receiver_id, transmitter_id, rssi, combinations_data,k))
        for timestamp, receiver_id, transmitter_id, rssi in all_combinations
    ]

    # 对新指标进行降序排序
    new_combinations.sort(key=lambda x: x[4], reverse=True)

    # 分配 transmitters 到 receivers
    used_transmitters = set()
    used_receivers = set()
    selected_combinations = []
    for timestamp, receiver_id, transmitter_id, rssi, new_metric in new_combinations:
        if transmitter_id not in used_transmitters and receiver_id not in used_receivers:
            selected_combinations.append((timestamp, receiver_id, transmitter_id, rssi))
            used_transmitters.add(transmitter_id)
            used_receivers.add(receiver_id)

    return selected_combinations

# 读取 normalized 数据的第 k 行和 combinations 数据的前 k-1 行 并更新数据
def update_combinations_with_kth_row(normalized_file, combinations_file, k):
    """
    读取 normalized 数据的第 k 行和 combinations 数据的前 k-1 行，并根据新的指标分配 transmitters 到 receivers。

    输入:
    - normalized_file: 归一化后的 Excel 文件名
    - combinations_file: 组合数据的 CSV 文件名
    - k: 要读取的行数 (从1开始)

    输出:
    - 无直接输出，将更新后的组合数据保存到 combinations_file 中
    """
    # 读取 normalized 数据的第 k 行和 combinations 数据的前 k-1 行
    normalized_data, combinations_data = read_kth_row_and_combinations(normalized_file, combinations_file, k)

    # 根据新的指标分配 transmitters 到 receivers
    new_combinations = allocate_transmitters_based_on_new_metric(normalized_data, combinations_data,k)

    # 读取现有的 combinations 数据
    existing_combinations = pd.read_csv(combinations_file, index_col=0)

    # 将新组合添加到现有的 combinations 数据中
    new_combinations_df = pd.DataFrame(new_combinations, columns=['Timestamp', 'Receiver_ID', 'Transmitter_ID', 'RSSI'])
    new_combinations_df.set_index('Timestamp', inplace=True)

    # 将新组合数据转换为行对应时间，列对应 Receiver 的格式
    pivot_new_combinations = new_combinations_df.pivot(columns='Receiver_ID', values='Transmitter_ID')

    # 更新组合数据
    updated_combinations = existing_combinations.combine_first(pivot_new_combinations)

    # 直接写入 CSV 文件中
    updated_combinations.to_csv(combinations_file, index=True)

    print(f"Updated combinations have been saved to {combinations_file}.")

def test_main_process():
    # 数据预处理
    preprocess_main()

    # 提问是否补全 receiver
    receiver_fill_input = input("预测时是否补全 receiver 编码？(这里会影响后续的连续窗口！！！) (y/n): ").strip().lower()
    receiver_fill = receiver_fill_input == 'y'

    normalized_file = 'dataNormalized.xlsx'
    combinations_file = 'combinations.csv'

    # 初始化 combinations.csv
    read_normalized_data_and_find_combinations(normalized_file, combinations_file, receiver_fill)

    # 从 k=2 开始不断迭代
    for k in range(2, 1000):  # 根据需要的 k 值范围调整
        update_combinations_with_kth_row(normalized_file, combinations_file, k)

# -----------------------------------------------------如果要看输入逻辑的话请阅读这里！！！！！！！！！！！！！
def run_main_process():
    # 启动 host.py
    host_process = subprocess.Popen(['python', 'host.py'])

    receiver_ids = input("请输入所有接收器编号 (用逗号分隔): ").strip().split(',')
    transmitter_ids = input("请输入所有发射器编号 (用逗号分隔): ").strip().split(',')

    columns = ['Timestamp'] + transmitter_ids
    combinations_df = pd.DataFrame(columns=columns)
    combinations_df.set_index('Timestamp', inplace=True)

    last_identifier = None
    last_positions = {}
    latest_updates = {}  # 用于存储每个发射器的最新位置和更新时间

    # 创建一个空的 CSV 文件，并写入表头
    write_to_csv('transmitter_changes.csv', [], columns)

    print("程序正在等待外部输入...")

    try:
        while True:
            with open('input.json', 'r') as f:
                input_data = f.read().strip()
                if input_data:
                    input_dict = json.loads(input_data)

                    if 'identifier' not in input_dict or 'inputs' not in input_dict or 'timestamp' not in input_dict:
                        print("输入数据格式错误，需要包含 'identifier', 'inputs' 和 'timestamp' 键.")
                        continue

                    identifier = input_dict['identifier']
                    inputs = input_dict['inputs']
                    timestamp = input_dict['timestamp']

                    if identifier == last_identifier:
                        #print("识别编码一致，跳过处理。")
                        continue
                    else:
                        last_identifier = identifier

                    # 更新后的输入逻辑
                    changes = []
                    for transmitter_id, receiver_id in inputs.items():
                        if transmitter_id not in columns:
                            print(f"发射器ID {transmitter_id} 不在初始发射器编号列表中.")
                            continue
                        if timestamp not in combinations_df.index:
                            combinations_df.loc[timestamp] = [np.nan] * len(combinations_df.columns)
                        combinations_df.at[timestamp, transmitter_id] = receiver_id

                        # 记录位置变化
                        if transmitter_id in last_positions:
                            if last_positions[transmitter_id] != receiver_id:
                                changes.append((timestamp, transmitter_id, receiver_id))
                                # 更新最新位置和时间
                                latest_updates[transmitter_id] = {"receiver_id": receiver_id, "timestamp": timestamp}
                        last_positions[transmitter_id] = receiver_id

                    if changes:
                        change_dict = {col: np.nan for col in columns}
                        change_dict['Timestamp'] = timestamp
                        for change in changes:
                            _, transmitter_id, receiver_id = change
                            change_dict[transmitter_id] = receiver_id
                            #print(f"时间: {timestamp}, 发射器: {transmitter_id}, 新接收器: {receiver_id}")

                        # 写入 CSV
                        write_to_csv('transmitter_changes.csv', [change_dict], columns)
                        # 按照ID升序排列最新位置和更新时间
                        # 按照ID升序排列最新位置和更新时间
                        sorted_latest_updates = {k: latest_updates[k] for k in
                                                 sorted(latest_updates, key=lambda x: int(x))}
                        # 打印最新位置和更新时间
                        print("最新位置和更新时间:", sorted_latest_updates)

                    #print(f"当前组合数据:\n{combinations_df}")
            time.sleep(1)  # 每1秒读取一次数据
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
    finally:
        # 关闭 host.py 进程
        host_process.terminate()

    print("running over")


def main():
    should_run = input("是否执行测试性能？(y/n)(y代表测试 n代表执行正式代码): ").strip().lower()
    if should_run == 'y':
        test_main_process()
    else:
        run_main_process()

if __name__ == "__main__":
    main()
